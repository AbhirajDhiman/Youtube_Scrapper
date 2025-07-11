
import os
import csv
import json
import io
import logging
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, session, make_response, flash, jsonify
from werkzeug.middleware.proxy_fix import ProxyFix
from youtube_service import YouTubeService
from database import DatabaseService
import openpyxl
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Create the app
app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "dev-secret-key-change-in-production")
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# Initialize services
youtube_service = YouTubeService()
db_service = DatabaseService()

@app.route('/')
def index():
    """Homepage with enhanced search form"""
    try:
        # Check if API service is available with better error handling
        if not youtube_service.api_key:
            quota_status = {
                'quota_used': 0, 
                'quota_limit': 10000, 
                'quota_remaining': 0, 
                'quota_percentage': 0, 
                'status': 'error',
                'quota_exceeded': True,
                'error_message': 'API key not configured. Please set GOOGLE_API_KEY environment variable.'
            }
            app.logger.warning("YouTube API key not found in environment variables")
        elif not youtube_service.youtube:
            quota_status = {
                'quota_used': 0, 
                'quota_limit': 10000, 
                'quota_remaining': 0, 
                'quota_percentage': 0, 
                'status': 'error',
                'quota_exceeded': True,
                'error_message': 'API service initialization failed. Check API key validity.'
            }
            app.logger.error("YouTube API service not initialized")
        else:
            try:
                quota_status = youtube_service.get_quota_status()
                # Ensure all required keys are present with proper structure
                required_keys = ['quota_used', 'quota_limit', 'quota_remaining', 'quota_percentage', 'status']
                for key in required_keys:
                    if key not in quota_status:
                        if key == 'quota_used':
                            quota_status[key] = getattr(youtube_service, 'daily_quota_used', 0)
                        elif key == 'quota_limit':
                            quota_status[key] = getattr(youtube_service, 'max_daily_quota', 10000)
                        elif key == 'quota_remaining':
                            used = getattr(youtube_service, 'daily_quota_used', 0)
                            limit = getattr(youtube_service, 'max_daily_quota', 10000)
                            quota_status[key] = max(0, limit - used)
                        elif key == 'quota_percentage':
                            used = getattr(youtube_service, 'daily_quota_used', 0)
                            limit = getattr(youtube_service, 'max_daily_quota', 10000)
                            quota_status[key] = (used / limit) * 100 if limit > 0 else 0
                        else:
                            quota_status[key] = 'healthy'

                app.logger.info(f"API Status: {quota_status}")
            except Exception as e:
                app.logger.error(f"Error getting quota status: {e}")
                quota_status = {
                    'quota_used': 0, 
                    'quota_limit': 10000, 
                    'quota_remaining': 10000, 
                    'quota_percentage': 0, 
                    'status': 'warning',
                    'quota_exceeded': False,
                    'error_message': f'Unable to check quota status: {str(e)}'
                }

    except Exception as e:
        app.logger.error(f"Failed to get quota status: {e}")
        quota_status = {
            'quota_used': 0, 
            'quota_limit': 10000, 
            'quota_remaining': 10000, 
            'quota_percentage': 0, 
            'status': 'warning',
            'quota_exceeded': False,
            'error_message': f'Unable to check quota status: {str(e)}'
        }

    return render_template('index.html', quota_status=quota_status)

@app.route('/search', methods=['POST'])
def search():
    """Enhanced search with advanced filtering and better error handling"""
    keyword = request.form.get('keyword', '').strip()

    if not keyword:
        flash('Please enter a search keyword', 'warning')
        return redirect(url_for('index'))

    # Validate keyword length
    if len(keyword) < 2:
        flash('Search keyword must be at least 2 characters long', 'warning')
        return redirect(url_for('index'))

    # Check if API key is available
    if not youtube_service.api_key:
        flash('YouTube API key not configured. Please contact administrator.', 'danger')
        return redirect(url_for('index'))

    # Check if YouTube service is initialized
    if not youtube_service.youtube:
        flash('YouTube API service not available. Please check API key configuration.', 'danger')
        return redirect(url_for('index'))

    # Get filter parameters
    filters = {}

    # Subscriber filters
    min_subs = request.form.get('min_subscribers')
    max_subs = request.form.get('max_subscribers')
    if min_subs:
        try:
            filters['min_subscribers'] = int(min_subs)
        except ValueError:
            flash('Invalid minimum subscribers value', 'warning')
    if max_subs:
        try:
            filters['max_subscribers'] = int(max_subs)
        except ValueError:
            flash('Invalid maximum subscribers value', 'warning')

    # Video count filters
    min_videos = request.form.get('min_videos')
    max_videos = request.form.get('max_videos')
    if min_videos:
        try:
            filters['min_videos'] = int(min_videos)
        except ValueError:
            flash('Invalid minimum videos value', 'warning')
    if max_videos:
        try:
            filters['max_videos'] = int(max_videos)
        except ValueError:
            flash('Invalid maximum videos value', 'warning')

    # Activity filters
    min_upload_freq = request.form.get('min_upload_frequency')
    max_days_since = request.form.get('max_days_since_upload')
    if min_upload_freq:
        try:
            filters['min_upload_frequency'] = float(min_upload_freq)
        except ValueError:
            flash('Invalid upload frequency value', 'warning')
    if max_days_since:
        try:
            filters['max_days_since_upload'] = int(max_days_since)
        except ValueError:
            flash('Invalid days since upload value', 'warning')

    # Get max results
    max_results = request.form.get('max_results', 25)  # Reduced default for speed
    try:
        max_results = min(int(max_results), 100)  # Reduced cap for performance
    except ValueError:
        max_results = 25

    app.logger.info(f"Searching for keyword: {keyword} with filters: {filters}")

    try:
        # Check quota before search
        quota_status = youtube_service.get_quota_status()
        if quota_status.get('quota_exceeded'):
            flash('Daily YouTube API quota exceeded. Please try again tomorrow.', 'warning')
            return redirect(url_for('index'))

        result = youtube_service.search_channels(keyword, max_results=max_results, filters=filters)

        if 'error' in result:
            app.logger.error(f"Search API error: {result['error']}")
            flash(f'Search error: {result["error"]}', 'danger')
            return redirect(url_for('index'))

        channels = result.get('channels', [])

        if not channels:
            flash('No channels found matching your criteria. Try different keywords or adjust your filters.', 'info')
            return redirect(url_for('index'))

        # Ensure all required keys are present
        required_keys = ['quota_used', 'quota_limit', 'quota_remaining', 'quota_percentage']
        for key in required_keys:
            if key not in result:
                if key == 'quota_used':
                    result[key] = youtube_service.daily_quota_used
                elif key == 'quota_limit':
                    result[key] = youtube_service.max_daily_quota
                elif key == 'quota_remaining':
                    result[key] = max(0, youtube_service.max_daily_quota - youtube_service.daily_quota_used)
                elif key == 'quota_percentage':
                    result[key] = (youtube_service.daily_quota_used / youtube_service.max_daily_quota) * 100

        # Store in session
        session['search_keyword'] = keyword
        session['search_results'] = channels
        session['search_filters'] = filters
        session['quota_status'] = result

        # Store in database if available
        if db_service.is_connected():
            try:
                db_service.save_search_history(keyword, len(channels), filters)
                for channel in channels:
                    db_service.save_channel_data(channel)
            except Exception as db_error:
                app.logger.warning(f"Database save error: {str(db_error)}")

        flash(f'Found {len(channels)} channels for "{keyword}"', 'success')
        return redirect(url_for('results'))

    except Exception as e:
        app.logger.error(f"Unexpected search error for keyword '{keyword}': {str(e)}")
        flash('Search temporarily unavailable. Please try again in a few minutes.', 'danger')
        return redirect(url_for('index'))

@app.route('/results')
def results():
    """Display search results with enhanced data"""
    channels = session.get('search_results', [])
    keyword = session.get('search_keyword', '')
    filters = session.get('search_filters', {})
    quota_status = session.get('quota_status', {})

    # Ensure quota_status has required keys
    if not quota_status or not all(key in quota_status for key in ['quota_used', 'quota_limit', 'quota_remaining', 'quota_percentage']):
        quota_status = {'quota_used': 0, 'quota_limit': 10000, 'quota_remaining': 10000, 'quota_percentage': 0, 'status': 'healthy'}

    if not channels:
        flash('No search results found. Please perform a search first.', 'warning')
        return redirect(url_for('index'))

    return render_template('results.html', 
                         channels=channels, 
                         keyword=keyword, 
                         filters=filters,
                         quota_status=quota_status,
                         total_results=len(channels))

@app.route('/export/<format>')
def export_data(format):
    """Export search results in multiple formats"""
    channels = session.get('search_results', [])
    keyword = session.get('search_keyword', 'search')

    if not channels:
        flash('No data to export. Please perform a search first.', 'warning')
        return redirect(url_for('index'))

    try:
        if format.lower() == 'csv':
            return export_csv(channels, keyword)
        elif format.lower() == 'excel':
            return export_excel(channels, keyword)
        elif format.lower() == 'json':
            return export_json(channels, keyword)
        else:
            flash('Invalid export format', 'danger')
            return redirect(url_for('results'))

    except Exception as e:
        app.logger.error(f"Export error for format '{format}': {str(e)}")
        flash(f'Error exporting data: {str(e)}', 'danger')
        return redirect(url_for('results'))

def export_csv(channels, keyword):
    """Export to CSV format"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Enhanced headers
    headers = [
        'Channel Name', 'URL', 'Subscribers', 'Video Count', 'Total Views',
        'Upload Frequency (per week)', 'Avg Views per Video', 'Avg Engagement Rate (%)',
        'Last Upload Date', 'Days Since Last Upload', 'Description', 'Country',
        'Website URL', 'Emails Found', 'Social Media', 'Contact Pages',
        'WHOIS Emails', 'Channel Age (days)', 'Custom URL'
    ]
    writer.writerow(headers)

    for channel in channels:
        quality = channel.get('quality_metrics', {})
        contact = channel.get('contact_info', {})

        # Calculate days since last upload
        days_since_upload = 'N/A'
        if quality.get('last_upload_date'):
            try:
                last_upload = quality['last_upload_date']
                if isinstance(last_upload, str):
                    last_upload = datetime.strptime(last_upload[:10], '%Y-%m-%d')
                days_since_upload = (datetime.now() - last_upload).days
            except:
                pass

        # Calculate channel age
        channel_age = 'N/A'
        if channel.get('created_at'):
            try:
                created_date = datetime.strptime(channel['created_at'][:10], '%Y-%m-%d')
                channel_age = (datetime.now() - created_date).days
            except:
                pass

        # Format social media links
        social_media = ', '.join([f"{platform}: {handle}" for platform, handle in contact.get('social_media', {}).items()])

        # Combine all emails
        all_emails = contact.get('emails', [])
        if contact.get('whois_info', {}).get('emails'):
            all_emails.extend(contact['whois_info']['emails'])
        all_emails = ', '.join(list(set(all_emails)))

        # Contact pages
        contact_pages = ', '.join(contact.get('website_contacts', {}).get('contact_pages', []))

        # WHOIS emails
        whois_emails = ', '.join(contact.get('whois_info', {}).get('emails', []))

        row = [
            channel.get('title', 'N/A'),
            channel.get('url', ''),
            channel.get('subscriber_count', 0),
            channel.get('video_count', 0),
            channel.get('view_count', 0),
            quality.get('upload_frequency', 'N/A'),
            quality.get('avg_views', 'N/A'),
            quality.get('avg_engagement_rate', 'N/A'),
            quality.get('last_upload_date', 'N/A'),
            days_since_upload,
            channel.get('description', '')[:200],  # Truncate description
            channel.get('country', 'Unknown'),
            channel.get('website_url', ''),
            all_emails,
            social_media,
            contact_pages,
            whois_emails,
            channel_age,
            channel.get('custom_url', '')
        ]
        writer.writerow(row)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=youtube_channels_{keyword.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

    return response

def export_excel(channels, keyword):
    """Export to Excel format with formatting"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "YouTube Channels"

    # Headers with formatting
    headers = [
        'Channel Name', 'URL', 'Subscribers', 'Video Count', 'Total Views',
        'Upload Frequency (per week)', 'Avg Views per Video', 'Avg Engagement Rate (%)',
        'Last Upload Date', 'Days Since Last Upload', 'Description', 'Country',
        'Website URL', 'Emails Found', 'Social Media', 'Contact Pages',
        'WHOIS Emails', 'Channel Age (days)', 'Custom URL'
    ]

    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add data
    for row_idx, channel in enumerate(channels, 2):
        quality = channel.get('quality_metrics', {})
        contact = channel.get('contact_info', {})

        # Calculate metrics
        days_since_upload = 'N/A'
        if quality.get('last_upload_date'):
            try:
                last_upload = quality['last_upload_date']
                if isinstance(last_upload, str):
                    last_upload = datetime.strptime(last_upload[:10], '%Y-%m-%d')
                days_since_upload = (datetime.now() - last_upload).days
            except:
                pass

        channel_age = 'N/A'
        if channel.get('created_at'):
            try:
                created_date = datetime.strptime(channel['created_at'][:10], '%Y-%m-%d')
                channel_age = (datetime.now() - created_date).days
            except:
                pass

        # Format data
        social_media = ', '.join([f"{platform}: {handle}" for platform, handle in contact.get('social_media', {}).items()])
        all_emails = contact.get('emails', [])
        if contact.get('whois_info', {}).get('emails'):
            all_emails.extend(contact['whois_info']['emails'])
        all_emails = ', '.join(list(set(all_emails)))
        contact_pages = ', '.join(contact.get('website_contacts', {}).get('contact_pages', []))
        whois_emails = ', '.join(contact.get('whois_info', {}).get('emails', []))

        data = [
            channel.get('title', 'N/A'),
            channel.get('url', ''),
            channel.get('subscriber_count', 0),
            channel.get('video_count', 0),
            channel.get('view_count', 0),
            quality.get('upload_frequency', 'N/A'),
            quality.get('avg_views', 'N/A'),
            quality.get('avg_engagement_rate', 'N/A'),
            quality.get('last_upload_date', 'N/A'),
            days_since_upload,
            channel.get('description', '')[:200],
            channel.get('country', 'Unknown'),
            channel.get('website_url', ''),
            all_emails,
            social_media,
            contact_pages,
            whois_emails,
            channel_age,
            channel.get('custom_url', '')
        ]

        for col, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col, value=value)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=youtube_channels_{keyword.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return response

def export_json(channels, keyword):
    """Export to JSON format"""
    # Add export timestamp and metadata
    export_data = {
        'export_info': {
            'keyword': keyword,
            'export_date': datetime.now().isoformat(),
            'total_channels': len(channels),
            'format': 'json'
        },
        'channels': channels
    }

    response = make_response(json.dumps(export_data, indent=2, default=str))
    response.headers['Content-Type'] = 'application/json'
    response.headers['Content-Disposition'] = f'attachment; filename=youtube_channels_{keyword.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'

    return response

@app.route('/history')
def history():
    """View search history"""
    if db_service.is_connected():
        search_history = db_service.get_search_history()
        return render_template('history.html', history=search_history, db_connected=True)
    else:
        # Use session-based history as fallback
        session_history = session.get('search_history', [])
        return render_template('history.html', history=session_history, db_connected=False)

@app.route('/stats')
def stats():
    """Database statistics and analytics"""
    if not db_service.is_connected():
        flash('Database not available. Statistics require MongoDB connection.', 'warning')
        return redirect(url_for('index'))

    try:
        stats_data = db_service.get_database_stats()
        return render_template('stats.html', stats=stats_data)
    except Exception as e:
        app.logger.error(f"Error loading stats: {str(e)}")
        flash('Error loading statistics', 'danger')
        return redirect(url_for('index'))

@app.route('/api/quota-status')
def api_quota_status():
    """API endpoint for quota status with error handling"""
    try:
        if not youtube_service.api_key:
            return jsonify({
                'quota_used': 0,
                'quota_limit': 10000,
                'quota_remaining': 0,
                'quota_percentage': 0,
                'status': 'error',
                'quota_exceeded': True,
                'error_message': 'API key not configured'
            })

        if not youtube_service.youtube:
            return jsonify({
                'quota_used': 0,
                'quota_limit': 10000,
                'quota_remaining': 0,
                'quota_percentage': 0,
                'status': 'error',
                'quota_exceeded': True,
                'error_message': 'API service not initialized'
            })

        status = youtube_service.get_quota_status()
        app.logger.info(f"Quota Status API Response: {status}")
        return jsonify(status)

    except Exception as e:
        app.logger.error(f"Error in quota status API: {str(e)}")
        return jsonify({
            'quota_used': 0,
            'quota_limit': 10000,
            'quota_remaining': 10000,
            'quota_percentage': 0,
            'status': 'error',
            'quota_exceeded': False,
            'error_message': f'API error: {str(e)}'
        }), 500

@app.route('/clear-results')
def clear_results():
    """Clear current search results"""
    session.pop('search_results', None)
    session.pop('search_keyword', None)
    session.pop('search_filters', None)
    session.pop('quota_status', None)
    flash('Search results cleared', 'info')
    return redirect(url_for('index'))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True, threaded=True)
